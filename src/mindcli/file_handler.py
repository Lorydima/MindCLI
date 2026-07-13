# File reading for attachments — supports text, PDF, DOCX, XLSX, images, and code files.

import os
from mindcli import state


# List of supported file extensions for attachment
SUPPORTED_EXTENSIONS = [".txt", ".pdf", ".docx", ".xlsx", ".xls", ".md", ".py", ".cpp", ".c", ".java", ".js", ".html", ".css", ".json", ".xml", ".csv", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"]


def read_file_content(file_path: str) -> str | None:
    """Reads a file and returns its text content. Supports PDF, DOCX, XLSX, images, and text files."""
    if not os.path.exists(file_path):
        state.console.print("[red]Error: File not found.[/red]")
        return None

    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext not in SUPPORTED_EXTENSIONS:
        state.console.print("[red]Error: File type not supported.[/red]")
        return None

    try:
        # Handle PDF files
        if file_path.lower().endswith(".pdf"):
            import pypdf
            reader = pypdf.PdfReader(file_path)
            text_list = []
            for page in reader.pages:
                text_list.append(page.extract_text())
            return "\n".join(t for t in text_list if t)

        # Handle Word documents
        elif file_path.lower().endswith(".docx"):
            from docx import Document
            doc = Document(file_path)
            text_list = []
            for paragraph in doc.paragraphs:
                text_list.append(paragraph.text)
            return "\n".join(text_list)

        # Handle Excel spreadsheets
        elif file_path.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            text_list = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_list.append(f"Sheet: {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    text_list.append(row_text)
            return "\n".join(text_list)

        # Handle image files (placeholder text only)
        elif file_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"]:
            return f"[Image file: {os.path.basename(file_path)}]"

        # Handle all text-based files (.txt, .md, code files, etc.)
        else:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()

    except Exception as e:
        state.console.print(f"[red]Error reading file: {e}[/red]")
        return None
