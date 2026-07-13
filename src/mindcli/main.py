# MindCLI V2.0 Source Code Date: 13/07/2026 Dev: LDM Dev.

'''
MindCLI is a command-line interface tool designed to facilitate interaction with AI offline models.

Git Hub Repository Link: "https://github.com/Lorydima/MindCLI"

MindCLI Website link: "https://lorydima.github.io/MindCLI/"

Before you use this code read the license in the LICENSE.txt or on Git Hub Repository.

If you discover a security vulnerability please read the file SECURITY.md on the Git Hub Repository.
'''

# Library for CLI Dev
from rich.console import Console, Group
from rich.panel import Panel
from rich.align import Align
from rich.rule import Rule
from rich.table import Table
from rich.text import Text
from rich.markdown import Markdown
from rich.live import Live
from rich.syntax import Syntax
from datetime import datetime
from docx import Document
from tavily import TavilyClient
import json
import os
import sys
import contextlib
import atexit
import pyperclip
import pyfiglet
import signal
import time
import webbrowser
import ollama
import pypdf
import openpyxl

# Get base path
def get_base_path():
    """Returns the base path for config and resources, adjusting for PyInstaller."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

# Console Setup
console = Console()

# Global Config State
config_model_path = None
config_base_prompt = None
config_user_name = None
config_parameters = {}

# Default AI Parameters 
DEFAULT_PARAMS = {
    "num_predict": 2048,
    "temperature": 0.5,
    "top_p": 0.9,
    "repeat_penalty": 1.1,
    "num_ctx": 4096
}

# Default base prompt 
DEFAULT_BASE_PROMPT = "You are a helpful coding assistant. Answer the user's questions clearly and provide code blocks when necessary."

# Chat history Init
chat_history = []

# Runtime State
active_model = None
active_base_prompt = None
ollama_ready = False
attached_file_content = None
attached_filename = None
agent_temp_content = None
config_download_source = "N/A"
config_device = "cpu"
config_tavily_api_key = ""
config_memories = []

# Ollama Process Managment
def ollama_is_available():
    """Returns True when the Ollama API is reachable."""
    try:
        ollama.list()
        return True
    except Exception:
        return False

def get_ollama_version() -> str:
    """Returns the installed Ollama version string if available."""
    try:
        if sys.platform == "win32":
            import subprocess
            result = subprocess.run(["ollama", "--version"], capture_output=True, text=True)
            output = result.stdout.strip()
        else:
            with os.popen("ollama --version") as pipe:
                output = pipe.read().strip()
        return output or "N/A"
    except Exception:
        return "N/A"

def start_ollama_process():
    """Starts Ollama in the background if it is not already reachable."""
    if ollama_is_available():
        return True

    try:
        if sys.platform == "win32":
            os.system("start /B ollama serve > NUL 2>&1")
        else:
            os.system("ollama serve > /dev/null 2>&1 &")
    except Exception:
        return False

    for _ in range(30):
        time.sleep(0.5)
        if ollama_is_available():
            return True

    return False

def kill_all_ollama_processes():
    """Best-effort kill for all Ollama processes."""
    try:
        if sys.platform == "win32":
            os.system("taskkill /IM ollama.exe /T /F > NUL 2>&1")
        else:
            os.system("pkill ollama > /dev/null 2>&1")
    except Exception:
        pass

def close_ollama_process():
    """Best-effort shutdown for an Ollama process started by this app."""
    kill_all_ollama_processes()

def shutdown_ollama_everywhere():
    """Stops Ollama processes and clears local state."""
    close_ollama_process()
    kill_all_ollama_processes()

def _handle_shutdown_signal(signum, frame):
    shutdown_ollama_everywhere()
    raise SystemExit(0)

for _sig in ("SIGINT", "SIGTERM", "SIGHUP"):
    if hasattr(signal, _sig):
        signal.signal(getattr(signal, _sig), _handle_shutdown_signal)

atexit.register(close_ollama_process)
atexit.register(kill_all_ollama_processes)

# Copy to clipboard function 
def copy_to_clipboard(text: str) -> bool:
    """Copies text using pyperclip."""
    try:
        pyperclip.copy(text)
        return True
    except Exception:
        pass
    return False

# Windows style masked prompt 
def prompt_masked_windows(prompt: str) -> str:
    """Prompts for hidden input using Windows-style asterisks."""
    try:
        import msvcrt
    except Exception:
        return console.input(prompt, password=True).strip()

    console.print(prompt, end="")
    chars = []

    while True:
        ch = msvcrt.getwch()
        if ch in ("\r", "\n"):
            sys.stdout.write("\n")
            sys.stdout.flush()
            break
        if ch == "\b":
            if chars:
                chars.pop()
                sys.stdout.write("\b \b")
                sys.stdout.flush()
            continue
        if ch in ("\x00", "\xe0"):
            msvcrt.getwch()
            continue
        chars.append(ch)
        sys.stdout.write("*")
        sys.stdout.flush()

    return "".join(chars).strip()


# Open Tavily site function 
def open_tavily_site():
    """Opens the Tavily API dashboard."""
    webbrowser.open("https://app.tavily.com/")


# Extract domain function 
def extract_domain(value: str) -> str:
    """Extracts a domain from a URL-like value."""
    if not value:
        return ""
    value = value.strip()
    if not value.startswith(("http://", "https://")):
        return ""
    try:
        from urllib.parse import urlparse as _urlparse
        return _urlparse(value).netloc.lower()
    except Exception:
        return ""

# Formati Taivily Context Funciton
def format_tavily_context(payload: dict) -> str:
    """Builds a compact context block from Tavily search results."""
    lines = []
    answer = (payload.get("answer") or "").strip()
    if answer:
        lines.append(f"Answer: {answer}")

    results = payload.get("results") or []
    for idx, result in enumerate(results[:5], 1):
        title = (result.get("title") or "").strip()
        url = (result.get("url") or "").strip()
        content = (result.get("content") or result.get("raw_content") or "").strip()
        lines.append(f"[{idx}] {title}")
        if url:
            lines.append(f"URL: {url}")
        if content:
            lines.append(f"Content: {content[:1800]}")
        lines.append("")

    return "\n".join(lines).strip()

# Generate ollama AI response function
def generate_ai_response(full_prompt: str) -> str | None:
    """Generates a response from Ollama and returns the cleaned text."""
    global active_model, chat_history

    ai_response = None

    options = {
        "num_predict": config_parameters.get("num_predict", 2048),
        "temperature": config_parameters.get("temperature", 0.5),
        "top_p": config_parameters.get("top_p", 0.9),
        "repeat_penalty": config_parameters.get("repeat_penalty", 1.1),
    }
    if "num_ctx" in config_parameters:
        options["num_ctx"] = config_parameters["num_ctx"]

    with console.status("[yellow]Thinking...[/yellow]", spinner="dots"):
        try:
            response = ollama.generate(
                model=active_model,
                prompt=full_prompt,
                options=options
            )

            raw = response.response if hasattr(response, 'response') else response.get('response', '')
            text = sanitize_response(str(raw or ""))

            prefixes_to_remove = ["AI:", "AI Assistant:", "Assistant:", "Response:"]
            for prefix in prefixes_to_remove:
                if text.startswith(prefix):
                    text = text[len(prefix):].strip()
                    break

            ai_response = text
            chat_history.append(f"{active_model} > {ai_response}")

        except Exception as e:
            print()
            console.print(f"[red]Error generating response: {str(e)}[/red]")

    return ai_response

# Tavily search function
def tavily_search(query: str, api_key: str) -> dict:
    """Calls Tavily Search through the official SDK and returns a dict payload."""
    client = TavilyClient(api_key=api_key)

    search_kwargs = {
        "query": query,
        "search_depth": "basic",
        "max_results": 5,
        "include_answer": True,
        "include_raw_content": True,
        "include_favicon": False,
    }

    domain = extract_domain(query)
    if domain:
        search_kwargs["include_domains"] = [domain]

    response = client.search(**search_kwargs)
    if isinstance(response, dict):
        return response

    payload = {}
    for key in ("answer", "query", "results", "images", "response_time"):
        if hasattr(response, key):
            payload[key] = getattr(response, key)
    return payload

# Check Ollama
def ensure_ollama_or_warn() -> bool:
    """Checks Ollama availability and prints a helpful message when missing."""
    if start_ollama_process():
        console.print("[green]Ollama is available.[/green]")
        return True

    console.print(
        "[red]Ollama not detected.[/red]\n"
        "[yellow]Download it here:[/yellow] https://ollama.com/download\n"
        "[yellow]Main site:[/yellow] https://ollama.com/"
    )
    return False

# Sanitize model filename function
def sanitize_model_filename(model_name: str) -> str:
    """Sanitizes model name to make it a safe JSON filename."""
    sanitized = (model_name or "").replace(":", "_").replace("/", "_").replace("\\", "_")
    return f"{sanitized}.json"


# Get model base name function
def get_model_base_name(model_name: str) -> str:
    """Returns the model name without the quantization tag."""
    if not model_name:
        return ""
    return model_name.split(":", 1)[0]


# Normalize parameters function
def normalize_parameters(params):
    """Normalizes stored parameters and backfills missing defaults."""
    normalized = DEFAULT_PARAMS.copy()
    if isinstance(params, dict):
        normalized.update(params)

    if "max_tokens" in normalized:
        normalized["num_predict"] = normalized.pop("max_tokens")
    if "temp" in normalized:
        normalized["temperature"] = normalized.pop("temp")

    normalized.setdefault("num_predict", DEFAULT_PARAMS["num_predict"])
    normalized.setdefault("temperature", DEFAULT_PARAMS["temperature"])
    normalized.setdefault("top_p", DEFAULT_PARAMS["top_p"])
    normalized.setdefault("repeat_penalty", DEFAULT_PARAMS["repeat_penalty"])
    normalized.setdefault("num_ctx", DEFAULT_PARAMS["num_ctx"])
    return normalized

# Open path with default app function
def open_path_with_default_app(path: str):
    """Opens a file or folder with the platform default application."""
    os.startfile(path)

# Get models directory function
def get_models_dir():
    """Returns the path to the model-specific configs directory, creating it if necessary."""
    d = os.path.join(get_base_path(), "models")
    os.makedirs(d, exist_ok=True)
    return d

# Open models folder content function
def open_models_folder_content():
    """Opens the models directory."""
    models_path = get_models_dir()

    try:
        if os.path.exists(models_path):
            os.startfile(models_path)
            console.print(f"[green]Opening models folder: {models_path}[/green]")
        else:
            console.print(f"[red]Models folder not found at: {models_path}[/red]")
    except Exception as e:
        console.print(f"[red]Error opening folder: {e}[/red]")

# Get chats directory function
def get_chats_dir():
    """Returns the path to the chats directory, creating it if necessary."""
    d = os.path.join(get_base_path(), "chats")
    os.makedirs(d, exist_ok=True)
    return d

# Open chats folder function
def open_chats_folder():
    """Opens the chats directory."""
    chats_path = get_chats_dir()
    try:
        if os.path.exists(chats_path):
            os.startfile(chats_path)
            console.print(f"[green]Opening chats folder: {chats_path}[/green]")
        else:
            console.print(f"[red]Chats folder not found at: {chats_path}[/red]")
    except Exception as e:
        console.print(f"[red]Error opening folder: {e}[/red]")

# Download model with progress function
def download_model_with_progress(model_name: str) -> bool:
    """Downloads an Ollama model displaying a simple spinner."""
    try:
        if not ollama_is_available():
            console.print("[red]Error: Ollama is not reachable.[/red]")
            return False

        with console.status(f"[yellow]Downloading model {model_name}...[/yellow]", spinner="dots"):
            for chunk in ollama.pull(model_name, stream=True):
                pass

        console.print(f"[yellow]Model {model_name} downloaded and ready to use[/yellow]")
        print()

        base_prompt = console.input("[cyan]Enter base prompt (press Enter for default) > [/cyan]").strip()
        if not base_prompt:
            base_prompt = "You are a helpful coding assistant. Answer the user's questions clearly and provide code blocks when necessary."
        console.print("[cyan]Enter parameters (press Enter for default values):[/cyan]")
        params = DEFAULT_PARAMS.copy()

        try:
            mt = console.input(f"[cyan]Num Predict (Current {params.get('num_predict', 2048)}) > [/cyan]").strip()
            if mt: params['num_predict'] = int(mt)

            tmp = console.input(f"[cyan]Temperature (Current {params.get('temperature', 0.5)}) > [/cyan]").strip()
            if tmp: params['temperature'] = float(tmp)

            tp = console.input(f"[cyan]Top P (Current {params.get('top_p', 0.9)}) > [/cyan]").strip()
            if tp: params['top_p'] = float(tp)

            rp = console.input(f"[cyan]Repeat Penalty (Current {params.get('repeat_penalty', 1.1)}) > [/cyan]").strip()
            if rp: params['repeat_penalty'] = float(rp)

            ctx = console.input(f"[cyan]Num Ctx (Current {params.get('num_ctx', 4096)}) > [/cyan]").strip()
            if ctx: params['num_ctx'] = int(ctx)
        except ValueError:
            console.print("[yellow]Invalid input, using default parameters.[/yellow]")
            params = DEFAULT_PARAMS.copy()
        print()
        console.print("[cyan]Select device for this model:[/cyan]")
        console.print(" 1. CPU (default)")
        console.print(" 2. GPU (NVIDIA)")
        device_choice = console.input("[cyan]Enter choice (1 or 2, press Enter for CPU) > [/cyan]").strip()
        device = "gpu" if device_choice == "2" else "cpu"
        
        print()
        if save_config_to_file(model_name, base_prompt, params, device=device):
            console.print("[green]Model configuration saved successfully![/green]")

        return True
    except Exception as e:
        console.print(f"[red]Error downloading model: {e}[/red]")
        return False

# Delete model Function
def delete_model_cmd():
    """Prints the list of installed models and lets the user choose one to delete."""
    try:
        if not ollama_is_available():
            console.print("[red]Error: Ollama is not reachable.[/red]")
            return

        with console.status("[cyan]Retrieving models list...[/cyan]", spinner="dots"):
            models_list = ollama.list()
            installed = [m.model for m in models_list.models if "-cloud" not in m.model.lower()]

        if not installed:
            console.print("[yellow]No local models found to delete.[/yellow]")
            return
        print()
        console.print("[bold yellow]Installed Local Models:[/bold yellow]")
        for m in installed:
            console.print(f" - [bold white]{m}[/bold white]")

        print()
        model_to_delete = console.input("[cyan]Enter the name of the model to delete > [/cyan]").strip()
        if not model_to_delete:
            console.print("[yellow]Cancelled.[/yellow]")
            return

        matched_model = None
        for m in installed:
            if m == model_to_delete or m.split(":")[0] == model_to_delete:
                matched_model = m
                break

        if not matched_model:
            console.print(f"[red]Error: Model '[bold]{model_to_delete}[/bold]' is not installed.[/red]")
            return

        confirm = console.input(f"[red]Are you sure you want to delete '[bold]{matched_model}[/bold]'? (y/n) > [/red]").strip().lower()
        if confirm == 'y':
            with console.status(f"[red]Deleting model '[bold]{matched_model}[/bold]'...[/red]", spinner="dots"):
                ollama.delete(model=matched_model)
            console.print("") 
            console.print(f"[green]Model '[bold]{matched_model}[/bold]' deleted successfully.[/green]")

            global config_model_path
            active_base = get_model_base_name(config_model_path)
            deleted_base = get_model_base_name(matched_model)
            if config_model_path == matched_model or active_base == deleted_base:
                fallback_model = next((m for m in installed if m != matched_model), "")
                if fallback_model:
                    console.print(f"[yellow]Warning: You deleted the active model. Switching to '[bold]{fallback_model}[/bold]'.[/yellow]")
                    config_model_path = fallback_model
                    save_config_to_file(config_model_path, config_base_prompt or DEFAULT_BASE_PROMPT)
                else:
                    console.print("[yellow]Warning: You deleted the last local model. Clearing the active model.[/yellow]")
                    config_model_path = ""
                    save_config_to_file("", config_base_prompt or DEFAULT_BASE_PROMPT)
        else:
            console.print("") 
            console.print("[yellow]Cancelled.[/yellow]")
    except Exception as e:
        console.print("") 
        console.print(f"[red]Error deleting model: {e}[/red]")

def list_models_cmd():
    """Displays a Rich table listing all installed local models and their sizes."""
    try:
        if not ollama_is_available():
            console.print("[red]Error: Ollama is not reachable.[/red]")
            return

        with console.status("[cyan]Retrieving models list...[/cyan]", spinner="dots"):
            models_list = ollama.list()
            installed = [m for m in models_list.models if "-cloud" not in m.model.lower()]

        if not installed:
            console.print("[yellow]No local models found.[/yellow]")
            return

        table = Table(title="Local Ollama Models List", border_style="cyan")
        table.add_column("Model Name", style="white bold")
        table.add_column("Weight (Size)", style="yellow")

        for m in installed:
            size_gb = f"{m.size / (1024**3):.2f} GB"
            table.add_row(m.model, size_gb)

        console.print(Align.center(table))
    except Exception as e:
        console.print(f"[red]Error listing models: {e}[/red]")

# suppress_stderr_fd
@contextlib.contextmanager
def suppress_stderr_fd():
    """Temporarily suppresses stderr to hide verbose model output."""
    try:
        devnull = open(os.devnull, "w")
        old_fd = os.dup(2)
        os.dup2(devnull.fileno(), 2)
        yield
    finally:
        try:
            os.dup2(old_fd, 2)
            os.close(old_fd)
        except Exception:
            pass
        try:
            devnull.close()
        except Exception:
            pass

# detect_gpu_device
def detect_gpu_device() -> str:
    """Detects if an NVIDIA GPU is available and returns 'gpu' or 'cpu'."""
    try:
        if sys.platform == "win32":
            result = os.system("nvidia-smi > NUL 2>&1")
        else:
            result = os.system("nvidia-smi > /dev/null 2>&1")
        if result == 0:
            return "gpu"
    except Exception:
        pass
    return "cpu"

# load_config
def load_config():
    """Loads active model and then loads its specific configuration JSON, creating defaults if missing."""
    global config_model_path, config_base_prompt, config_parameters, config_download_source, config_device, config_tavily_api_key

    config_dir = os.path.join(get_base_path(), "configs")
    os.makedirs(config_dir, exist_ok=True)

    paths_path = os.path.join(config_dir, "paths.json")
    hardware_path = os.path.join(config_dir, "hardware.json")
    tavily_path = os.path.join(config_dir, "tavily_API.json")

    saved_model = ""
    saved_device = detect_gpu_device()
    config_tavily_api_key = ""

    if os.path.exists(paths_path):
        try:
            with open(paths_path, "r", encoding="utf-8") as f:
                pcfg = json.load(f)
                saved_model = pcfg.get("active_model", "")
        except Exception:
            pass

    if os.path.exists(tavily_path):
        try:
            with open(tavily_path, "r", encoding="utf-8") as f:
                tcfg = json.load(f)
                config_tavily_api_key = tcfg.get("tavily_api_key", "")
        except Exception:
            pass

    if os.path.exists(hardware_path):
        try:
            with open(hardware_path, "r", encoding="utf-8") as f:
                hcfg = json.load(f)
                saved_device = hcfg.get("device", saved_device)
        except Exception:
            pass

    desktop_default = os.path.join(os.path.expanduser("~"), "Desktop")

    if not os.path.exists(paths_path):
        try:
            with open(paths_path, "w", encoding="utf-8") as f:
                json.dump({"active_model": "", "desktop_path": desktop_default}, f, indent=2)
        except Exception:
            pass

    if not os.path.exists(tavily_path):
        try:
            with open(tavily_path, "w", encoding="utf-8") as f:
                json.dump({"tavily_api_key": ""}, f, indent=2)
        except Exception:
            pass

    if not os.path.exists(hardware_path):
        try:
            with open(hardware_path, "w", encoding="utf-8") as f:
                json.dump({"device": saved_device}, f, indent=2)
        except Exception:
            pass

    config_model_path = saved_model
    config_device = saved_device

    config_base_prompt = DEFAULT_BASE_PROMPT
    config_download_source = "N/A"
    config_parameters = DEFAULT_PARAMS.copy()

    if saved_model:
        models_dir = get_models_dir()
        model_cfg_filename = sanitize_model_filename(saved_model)
        model_cfg_path = os.path.join(models_dir, model_cfg_filename)

        model_config = {}
        if os.path.exists(model_cfg_path):
            try:
                with open(model_cfg_path, "r", encoding="utf-8") as f:
                    model_config = json.load(f)
            except Exception:
                pass

        if model_config.get("base_prompt"):
            config_base_prompt = model_config["base_prompt"]

        config_download_source = model_config.get("download_source", "N/A")
        config_device = model_config.get("device", saved_device)
        config_parameters = normalize_parameters(model_config.get("parameters", DEFAULT_PARAMS.copy()))

    load_memories()

# save_config_to_file   
def save_config_to_file(model_name, base_prompt, parameters=None, download_source=None, device=None):
    """Saves the configuration for a specific model to its individual JSON file, and updates active_model."""
    global config_parameters, config_download_source, config_device, config_model_path, config_tavily_api_key

    if parameters is None:
        parameters = config_parameters
    if download_source is None:
        download_source = config_download_source
    if device is None:
        device = config_device
    if not base_prompt:
        base_prompt = DEFAULT_BASE_PROMPT

    config_dir = os.path.join(get_base_path(), "configs")
    os.makedirs(config_dir, exist_ok=True)

    paths_path = os.path.join(config_dir, "paths.json")
    tavily_path = os.path.join(config_dir, "tavily_API.json")
    hardware_path = os.path.join(config_dir, "hardware.json")

    try:
        pcfg = {}
        if os.path.exists(paths_path):
            with open(paths_path, "r", encoding="utf-8") as f:
                pcfg = json.load(f)
        pcfg["active_model"] = model_name
        with open(paths_path, "w", encoding="utf-8") as f:
            json.dump(pcfg, f, indent=2)
    except Exception as e:
        console.print(f"[red]Error saving paths: {e}[/red]")

    try:
        tcfg = {}
        if os.path.exists(tavily_path):
            with open(tavily_path, "r", encoding="utf-8") as f:
                tcfg = json.load(f)
        tcfg["tavily_api_key"] = config_tavily_api_key
        with open(tavily_path, "w", encoding="utf-8") as f:
            json.dump(tcfg, f, indent=2)
    except Exception as e:
        console.print(f"[red]Error saving Tavily API key: {e}[/red]")

    try:
        hcfg = {}
        if os.path.exists(hardware_path):
            with open(hardware_path, "r", encoding="utf-8") as f:
                hcfg = json.load(f)
        hcfg["device"] = device
        with open(hardware_path, "w", encoding="utf-8") as f:
            json.dump(hcfg, f, indent=2)
    except Exception as e:
        console.print(f"[red]Error saving hardware config: {e}[/red]")

    config_model_path = model_name
    config_device = device

    if not model_name:
        config_parameters = normalize_parameters(parameters)
        config_download_source = download_source
        config_device = device
        return True

    models_dir = get_models_dir()
    model_cfg_filename = sanitize_model_filename(model_name)
    model_cfg_path = os.path.join(models_dir, model_cfg_filename)

    cfg = {
        "base_prompt": base_prompt,
        "download_source": download_source,
        "device": device,
        "parameters": normalize_parameters(parameters)
    }

    try:
        with open(model_cfg_path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        config_parameters = normalize_parameters(parameters)
        config_download_source = download_source
        config_device = device
        return True
    except Exception as e:
        console.print(f"[red]Error saving model config: {e}[/red]")
        return False


# sanitize_response
def sanitize_response(text: str) -> str:
    """Sanitizes text output from the AI response."""
    if not text:
        return ""
    t = text.strip()
    return t.strip()

# typing_effect
def typing_effect(text: str, model_name: str):
    """Displays a typing effect with Rich while preserving code fences and markdown."""
    if not text:
        return

    console.print(Text(f"{model_name} >", style="bold cyan"))

    parts = text.split("```")
    for index, part in enumerate(parts):
        if index % 2 == 0:
            cleaned = part.strip()
            if not cleaned:
                continue

            buffer = ""
            with Live(Markdown(""), console=console, refresh_per_second=30) as live:
                for ch in cleaned:
                    buffer += ch
                    try:
                        live.update(Markdown(buffer))
                    except Exception:
                        live.update(Text(buffer))
                    time.sleep(0.006)
            console.print()
        else:
            lines = part.splitlines()
            language = lines[0].strip() if lines else ""
            code = "\n".join(lines[1:]) if len(lines) > 1 else ""
            if code.strip():
                console.print(Syntax(code, language or "text", theme="monokai", line_numbers=False))

# open_models_folder
def open_models_folder():
    """Opens the local Ollama models directory."""
    models_path = os.environ.get("OLLAMA_MODELS")
    if not models_path:
        home = os.path.expanduser("~")
        if sys.platform == "win32":
            models_path = os.path.join(home, ".ollama", "models")
        elif sys.platform == "darwin":
            models_path = os.path.join(home, ".ollama", "models")
        else: 
            system_path = "/usr/share/ollama/.ollama/models"
            user_path = os.path.join(home, ".ollama", "models")
            if os.path.exists(system_path):
                models_path = system_path
            else:
                models_path = user_path

    if not os.path.exists(models_path):
        try:
            os.makedirs(models_path, exist_ok=True)
        except Exception:
            pass

    try:
        if os.path.exists(models_path):
            os.startfile(models_path)
            console.print(f"[green]Opening Ollama models folder: {models_path}[/green]")
        else:
            console.print(f"[red]Ollama models folder not found at: {models_path}[/red]")
    except Exception as e:
        console.print(f"[red]Error opening folder: {e}[/red]")


# introduction
def introduction():
    """Displays the main banner and introduction of the app."""
    banner = pyfiglet.figlet_format("MindCLI")
    version_text = Text("\nVersion 2.0", style="bold", justify="center")
    banner_content = Group(
        banner.rstrip(),
        version_text
    )

    banner_panel = Panel(banner_content, style="red", expand=False)

    console.print(Align.center(banner_panel))
    console.print(Rule(style="white"))

    start_ollama_process()
    if not ollama_is_available():
        console.print(
            "[red]Ollama not detected.[/red]\n"
            "[yellow]Download it here:[/yellow] https://ollama.com/download\n"
            "[yellow]Main site:[/yellow] https://ollama.com/"
        )
    load_config()
    command_list_function()


# agent_mode_function
def agent_mode_function():
    """Agent mode for working directly with files."""
    global agent_temp_content, chat_history

    agent_temp_content = None
    warning_agent = Text("WARNING: Review AI-generated file content carefully. If executing code, verify it is safe before use.", style="red bold")
    console.print(Align.center("[bold cyan]Agent Mode[/bold cyan]"))
    console.print(Align.center("[yellow]In this mode, the AI can work directly with files on your PC.[/yellow]"))
    chat_history.append("USER: Agent mode activated")
    while True:
        print()
        console.print("[cyan]Do you want to [bold]create[/bold] a new file or [bold]edit[/bold] an existing one?[/cyan]")
        agent_input = console.input("[magenta]Agent (create/edit/exit) > [/magenta]").strip().lower()

        if agent_input == "exit":
            console.print("[bold cyan]Exiting Agent Mode[/bold cyan]")
            break

        if agent_input in ("create", "new"):
            chat_history.append("Agent: Create File")
            filename = console.input("[cyan]Enter filename with extension (e.g., script.py, document.txt) > [/cyan]").strip()
            if not filename:
                console.print("[red]Filename cannot be empty.[/red]")
                continue

            supported_extensions = [".txt", ".md", ".py", ".cpp", ".c", ".java", ".js", ".html", ".css", ".json", ".xml"]
            file_ext = os.path.splitext(filename)[1].lower()
            if file_ext not in supported_extensions:
                console.print(f"[red]File type not supported. Supported types: {', '.join(supported_extensions)}[/red]")
                continue

            prompt = console.input("[cyan]Enter prompt for the AI > [/cyan]").strip()
            if not prompt:
                console.print("[red]Prompt cannot be empty.[/red]")
                continue

            chat_history.append(f"User: {filename}")
            chat_history.append(f"{active_model} prompt > {prompt}")

            refinement_context = ""
            if agent_temp_content:
                refinement_context = f"\n\nPrevious unsaved version to refine:\n{agent_temp_content}\n\nRefine it based on the new request."
                console.print("[yellow]Including previous unsaved content for refinement.[/yellow]")

            ai_prompt = f"Create content for a file named '{filename}' based on this request: {prompt}{refinement_context}\n\nRespond ONLY with the file content, without any additional text, explanations, or formatting."
            ai_response = generate_ai_response(ai_prompt)

            if ai_response:
                cleaned = ai_response
                prefixes = ["ecco", "certo", "ecco qui", "certo ecco", "ecco il file", "certo, ecco", "certo! ecco"]
                for p in prefixes:
                    if cleaned.lower().startswith(p):
                        cleaned = cleaned[len(p):].strip().lstrip(":,!.- ")
                        break

                chat_history.append(f"{active_model} agent response > {cleaned}")

                print()
                console.print(Align.center(Text("AI Generated Content", style="bold yellow")))
                console.print(Panel(cleaned, title=filename, border_style="yellow"))

                approval = console.input("[cyan]Do you want to save this content? (y/n) > [/cyan]").strip().lower()
                if approval == 'y':
                    chat_history.append("User: Content accepted and saved")
                    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                    file_path = os.path.join(desktop_path, filename)

                    try:
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(cleaned)
                        console.print(f"[green]File '[bold]{filename}[/bold]' created successfully on Desktop![/green]")
                        agent_temp_content = None
                    except Exception as e:
                        console.print(f"[red]Error creating file: {e}[/red]")
                else:
                    chat_history.append("User: Content not accepted")
                    console.print("[yellow]Content not saved. It will be included in the next prompt for refinement.[/yellow]")
                    agent_temp_content = cleaned
            continue

        if agent_input == "edit":
            chat_history.append("Agent: Edit File")
            file_path = console.input("[cyan]Enter the full path of the file to edit > [/cyan]").strip()
            if not file_path:
                console.print("[red]File path cannot be empty.[/red]")
                continue

            if not os.path.exists(file_path):
                console.print("[red]File not found.[/red]")
                continue

            supported_extensions = [".txt", ".md", ".py", ".cpp", ".c", ".java", ".js", ".html", ".css", ".json", ".xml"]
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in supported_extensions:
                console.print(f"[red]File type not supported for editing. Supported types: {', '.join(supported_extensions)}[/red]")
                continue

            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    current_content = f.read()
            except Exception as e:
                console.print(f"[red]Error reading file: {e}[/red]")
                continue

            console.print(Align.center(Text(f"Current File Content: {os.path.basename(file_path)}", style="bold white")))

            prompt = console.input("[cyan]Enter prompt for the AI (describe what changes to make) > [/cyan]").strip()
            if not prompt:
                console.print("[red]Prompt cannot be empty.[/red]")
                continue

            chat_history.append(f"User: {file_path}")
            chat_history.append(f"{active_model} prompt > {prompt}")

            refinement_context = ""
            if agent_temp_content:
                refinement_context = f"\n\nPrevious unsaved version to refine:\n{agent_temp_content}\n\nRefine it based on the new request."
                console.print("[yellow]Including previous unsaved content for refinement.[/yellow]")

            ai_prompt = f"Current content of {os.path.basename(file_path)}:\n{current_content}\n\nRequest: {prompt}{refinement_context}\n\nRespond ONLY with the modified file content, without any additional text, explanations, or formatting."
            ai_response = generate_ai_response(ai_prompt)

            if ai_response:
                cleaned = ai_response
                prefixes = ["ecco", "certo", "ecco qui", "certo ecco", "ecco il file", "certo, ecco", "certo! ecco", "ecco il contenuto"]
                for p in prefixes:
                    if cleaned.lower().startswith(p):
                        cleaned = cleaned[len(p):].strip().lstrip(":,!.- ")
                        break

                chat_history.append(f"{active_model} agent response > {cleaned}")

                print()
                console.print(Align.center(Text("AI Generated Content", style="bold yellow")))
                console.print(Panel(cleaned, title=os.path.basename(file_path), border_style="yellow"))

                approval = console.input("[cyan]Do you want to save these changes? (y/n) > [/cyan]").strip().lower()
                if approval == 'y':
                    chat_history.append("User: Changes accepted and saved")
                    try:
                        with open(file_path, "w", encoding="utf-8") as f:
                            f.write(cleaned)
                        console.print(f"[green]File '[bold]{os.path.basename(file_path)}[/bold]' updated successfully![/green]")
                        agent_temp_content = None
                    except Exception as e:
                        console.print(f"[red]Error updating file: {e}[/red]")
                else:
                    chat_history.append("User: Changes not accepted")
                    console.print("[yellow]Changes not saved. They will be included in the next prompt for refinement.[/yellow]")
                    agent_temp_content = cleaned
            continue

        console.print("[yellow]Unknown command. Type 'create', 'edit', or 'exit'.[/yellow]")


# change_model_function
def change_model_function():
    """Change the currently active AI model during chat."""
    global active_model, active_base_prompt, ollama_ready, attached_file_content, attached_filename, chat_history
    global config_model_path, config_base_prompt, config_parameters, config_device

    try:
        with console.status("[cyan]Retrieving models list...[/cyan]", spinner="dots"):
            models_list = ollama.list()
            installed = [m.model for m in models_list.models if "-cloud" not in m.model.lower()]

        if not installed:
            console.print("[yellow]No local models found. Please download a model first.[/yellow]")
            return

        console.print("[bold yellow]Available Local Models:[/bold yellow]")
        console.print("")
        for i, m in enumerate(installed, 1):
            console.print(f" [bold white]{i}. {m}[/bold white]")

        print()
        choice = console.input("[cyan]Enter the number or name of the model to switch to > [/cyan]").strip()

        if not choice:
            console.print("[yellow]Cancelled.[/yellow]")
            return

        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(installed):
                model_name = installed[idx]
            else:
                console.print("[red]Invalid selection.[/red]")
                return
        else:
            model_name = None
            for m in installed:
                if m == choice or m.split(":")[0] == choice:
                    model_name = m
                    break
            if not model_name:
                console.print(f"[red]Model '[bold]{choice}[/bold]' not found.[/red]")
                return

        models_dir = get_models_dir()
        model_cfg_filename = sanitize_model_filename(model_name)
        model_cfg_path = os.path.join(models_dir, model_cfg_filename)

        model_config = {}
        if os.path.exists(model_cfg_path):
            try:
                with open(model_cfg_path, "r", encoding="utf-8") as f:
                    model_config = json.load(f)
            except Exception as e:
                console.print(f"[yellow]Warning: Could not load config for {model_name}: {e}[/yellow]")

        config_model_path = model_name
        config_base_prompt = model_config.get("base_prompt", DEFAULT_BASE_PROMPT)
        config_parameters = normalize_parameters(model_config.get("parameters", DEFAULT_PARAMS.copy()))
        config_device = model_config.get("device", config_device)

        active_model = model_name
        active_base_prompt = config_base_prompt
        chat_history = []
        attached_file_content = None
        attached_filename = None

        save_config_to_file(model_name, config_base_prompt, config_parameters, device=config_device)

        console.print(f"[green]Model switched to '[bold]{active_model}[/bold]' successfully! Chat history cleared.[/green]")

    except Exception as e:
        console.print(f"[red]Error changing model: {str(e)}[/red]")
        return


# load_function
def load_function():
    """Loads the Ollama AI model and starts the chat loop."""
    global active_model, active_base_prompt, ollama_ready, attached_file_content, attached_filename

    active_model = None
    active_base_prompt = None
    ollama_ready = False
    attached_file_content = None
    attached_filename = None

    try:
        with console.status("[cyan]Retrieving models list...[/cyan]", spinner="dots"):
            models_list = ollama.list()
            installed = [m.model for m in models_list.models if "-cloud" not in m.model.lower()]

        if not installed:
            console.print("[yellow]No local models found. Please download a model first.[/yellow]")
            return

        console.print("[bold yellow]Available Local Models:[/bold yellow]")
        console.print("") 
        for i, m in enumerate(installed, 1):
            console.print(f" [bold white]{i}. {m}[/bold white]")

        print()
        choice = console.input("[cyan]Enter the number or name of the model to run > [/cyan]").strip()

        if not choice:
            console.print("[yellow]Cancelled.[/yellow]")
            return

        if choice.isdigit():
            idx = int(choice) - 1
            if 0 <= idx < len(installed):
                model_name = installed[idx]
            else:
                console.print("[red]Invalid selection.[/red]")
                return
        else:
            model_name = None
            for m in installed:
                if m == choice or m.split(":")[0] == choice:
                    model_name = m
                    break
            if not model_name:
                console.print(f"[red]Model '[bold]{choice}[/bold]' not found.[/red]")
                return

        models_dir = get_models_dir()
        model_cfg_filename = sanitize_model_filename(model_name)
        model_cfg_path = os.path.join(models_dir, model_cfg_filename)

        model_config = {}
        if os.path.exists(model_cfg_path):
            try:
                with open(model_cfg_path, "r", encoding="utf-8") as f:
                    model_config = json.load(f)
            except Exception as e:
                console.print(f"[yellow]Warning: Could not load config for {model_name}: {e}[/yellow]")

        global config_model_path, config_base_prompt, config_parameters, config_device
        config_model_path = model_name
        config_base_prompt = model_config.get("base_prompt", DEFAULT_BASE_PROMPT)
        config_parameters = normalize_parameters(model_config.get("parameters", DEFAULT_PARAMS.copy()))
        config_device = model_config.get("device", config_device)

        active_model = model_name
        active_base_prompt = config_base_prompt

        with console.status("[green]Starting Ollama process...[/green]", spinner="dots"):
            started = ensure_ollama_or_warn()

        if not started:
            return

        ollama_ready = True
        console.print(f"[green]Model '[bold]{active_model}[/bold]' is active and ready.[/green]")

        chat_command_list_function()
        chat_loop()

    except Exception as e:
        console.print(f"[red]Error during startup: {str(e)}[/red]")
        return


# command_list_function
def command_list_function():
    """Main menu command loop."""
    global config_model_path, config_base_prompt, config_tavily_api_key

    console.print(
        Panel(
            "[cyan]More[/cyan] - view more commands.\n"
            "[green]Run[/green] - initialize ollama and start the chat.\n"
            "[red]Exit[/red] - close the app.",
            title="Main Command",
            border_style="yellow"
        )
    )
    console.print(Rule(style="white"))

    while True:
        print()
        command = console.input("[cyan]Command > [/cyan]").strip().lower()

        if command == "run":
            load_function()

        elif command == "exit":
            shutdown_ollama_everywhere()
            sys.exit()

        elif command == "":
            ensure_ollama_or_warn()

        elif command in ("more", "more commands"):
            print()
            console.print("[bold white]Models Edit[/bold white]")
            print()
            console.print(" • [cyan]Edit Base Prompt[/cyan] - modify the base prompt used by the AI.")
            console.print(" • [cyan]Change Parameters[/cyan] - modify model parameters (tokens, temp, etc.)")
            print()
            console.print("[bold white]Models Management[/bold white]")
            print()
            console.print(" • [cyan]Download[/cyan] - download models.")
            console.print(" • [cyan]Delete[/cyan] - remove a model.")
            console.print(" • [cyan]List[/cyan] - view installed models.")
            print()
            console.print("[bold white]Tavily Web search Commands[/bold white]")
            print()
            console.print(" • [cyan]API Config[/cyan] - configure the Tavily API key.")
            console.print(" • [cyan]API Clear[/cyan] - remove the saved Tavily API key.")
            print()
            console.print("[bold white]Chats Commands[/bold white]")
            print()
            console.print(" • [cyan]Chats List[/cyan] - list all saved chat files.")
            console.print(" • [cyan]Open Chat[/cyan] - display a saved chat in the terminal.")
            console.print(" • [cyan]Remove Chat[/cyan] - delete a saved chat file.")
            print()
            console.print("[bold white]Memory Management[/bold white]")
            print()
            console.print(" • [cyan]Memory Add[/cyan] - add a memory for the AI.")
            console.print(" • [cyan]Memory View[/cyan] - view all saved memories.")
            console.print(" • [cyan]Memory Remove[/cyan] - remove a specific memory.")
            console.print(" • [cyan]Memory Clear[/cyan] - clear all memories.")
            print()
            console.print("[bold white]App Commands[/bold white]")
            print()
            console.print(" • [cyan]Help[/cyan] - open the help file.")
            console.print(" • [cyan]License[/cyan] - view the license.")
            console.print(" • [cyan]Info[/cyan] - view app information.")
            print()
            continue

        elif command == "download":
            model_to_download = console.input("[cyan]Enter model name to download > [/cyan]").strip()
            if model_to_download:
                result = download_model_with_progress(model_to_download)
                if result:
                    console.print(f"[green]Model {model_to_download} downloaded[/green]")
            else:
                console.print("[yellow]Cancelled.[/yellow]")

        elif command == "delete":
            delete_model_cmd()

        elif command == "list":
            list_models_cmd()

        elif command == "info":
            ollama_version = get_ollama_version()
            info_text = (
                "Developer: LDM Dev\n"
                "App Version: 2.0\n"
                f"{ollama_version}\n"
                "Repository GitHub: https://github.com/Lorydima/MindCLI\n"
                "Website: https://Lorydima.github.io/MindCLI/\n"
                "License: GPL "
            )
            console.print(Panel(info_text, title="App Information", border_style="cyan"))

        elif command in ("folder", "models folder"):
            open_models_folder()

        elif command == "models folder":
            open_models_folder_content()

        elif command == "chats list":
            list_chats_cmd()

        elif command == "open chat":
            open_chat_cmd()

        elif command == "remove chat":
            remove_chat_cmd()

        elif command == "memory add":
            memory_add()

        elif command == "memory view":
            memory_view()

        elif command == "memory remove":
            memory_remove()

        elif command == "memory clear":
            memory_clear()

        elif command == "api config":
            open_tavily_site()
            print()
            key = prompt_masked_windows("[cyan]Enter Tavily API key > [/cyan]")
            if key:
                config_tavily_api_key = key
                save_config_to_file(config_model_path or "", config_base_prompt or DEFAULT_BASE_PROMPT)
                console.print("[green]Tavily API key saved.[/green]")
            else:
                console.print("[yellow]Operation cancelled.[/yellow]")

        elif command == "api clear":
            config_tavily_api_key = ""
            save_config_to_file(config_model_path or "", config_base_prompt or DEFAULT_BASE_PROMPT)
            console.print("[yellow]Tavily API key cleared.[/yellow]")

        elif command == "help":
            help_path = os.path.join(get_base_path(), "docs", "MindCLI_UserGuide.pdf")
            try:
                if os.path.exists(help_path):
                    os.startfile(help_path)
                    console.print("[green]Opening MindCLI_UserGuide.pdf[/green]")
                else:
                    console.print(f"[red]MindCLI_UserGuide.pdf not found in the docs folder[/red]")
            except Exception as e:
                console.print(f"[red]Error opening help file: {e}[/red]")

        elif command == "license":
            license_path = os.path.join(get_base_path(), "docs", "LICENSE.txt")
            try:
                if os.path.exists(license_path):
                    open_path_with_default_app(license_path)
                    console.print("[green]Opening LICENSE.txt[/green]")
                else:
                    console.print(f"[red]LICENSE.txt not found in the docs folder[/red]")
            except Exception as e:
                console.print(f"[red]Error opening license: {e}[/red]")

        elif command == "edit base prompt":
            model_name = console.input("[cyan]Enter model name to edit base prompt > [/cyan]").strip()
            if not model_name:
                console.print("[yellow]Operation cancelled.[/yellow]")
                continue
            models_dir = get_models_dir()
            model_cfg_filename = sanitize_model_filename(model_name)
            model_cfg_path = os.path.join(models_dir, model_cfg_filename)

            current_prompt = "You are a helpful coding assistant. Answer the user's questions clearly and provide code blocks when necessary."
            if os.path.exists(model_cfg_path):
                try:
                    with open(model_cfg_path, "r", encoding="utf-8") as f:
                        m_cfg = json.load(f)
                        current_prompt = m_cfg.get("base_prompt", current_prompt)
                except Exception:
                    pass

            console.print(f"[cyan]Current prompt: {current_prompt}[/cyan]")
            new_prompt = console.input("[cyan]Enter new base prompt > [/cyan]").strip()
            if new_prompt:
                current_params = DEFAULT_PARAMS.copy()
                if os.path.exists(model_cfg_path):
                    try:
                        with open(model_cfg_path, "r", encoding="utf-8") as f:
                            m_cfg = json.load(f)
                            current_params = m_cfg.get("parameters", current_params)
                    except Exception:
                        pass

                current_device = "cpu"
                if os.path.exists(model_cfg_path):
                    try:
                        with open(model_cfg_path, "r", encoding="utf-8") as f:
                            m_cfg = json.load(f)
                            current_device = m_cfg.get("device", current_device)
                    except Exception:
                        pass

                if save_config_to_file(model_name, new_prompt, current_params, device=current_device):
                    console.print("[green]Base prompt updated![/green]")
            else:
                console.print("[yellow]Operation cancelled.[/yellow]")

        elif command == "change parameters":
            model_name = console.input("[cyan]Enter model name to change parameters > [/cyan]").strip()
            if not model_name:
                console.print("[yellow]Operation cancelled.[/yellow]")
                continue
            models_dir = get_models_dir()
            model_cfg_filename = sanitize_model_filename(model_name)
            model_cfg_path = os.path.join(models_dir, model_cfg_filename)

            current_params = DEFAULT_PARAMS.copy()
            if os.path.exists(model_cfg_path):
                try:
                    with open(model_cfg_path, "r", encoding="utf-8") as f:
                        m_cfg = json.load(f)
                        current_params = m_cfg.get("parameters", current_params)
                except Exception:
                    pass

            console.print("[cyan]Enter new parameters (press Enter for current/default values):[/cyan]")
            params = current_params.copy()
            try:
                mt = console.input(f"[cyan]Num Predict (Current {params.get('num_predict', 2048)}) > [/cyan]").strip()
                if mt: params['num_predict'] = int(mt)

                tmp = console.input(f"[cyan]Temperature (Current {params.get('temperature', 0.5)}) > [/cyan]").strip()
                if tmp: params['temperature'] = float(tmp)

                tp = console.input(f"[cyan]Top P (Current {params.get('top_p', 0.9)}) > [/cyan]").strip()
                if tp: params['top_p'] = float(tp)

                rp = console.input(f"[cyan]Repeat Penalty (Current {params.get('repeat_penalty', 1.1)}) > [/cyan]").strip()
                if rp: params['repeat_penalty'] = float(rp)

                ctx = console.input(f"[cyan]Num Ctx (Current {params.get('num_ctx', 4096)}) > [/cyan]").strip()
                if ctx: params['num_ctx'] = int(ctx)

                current_device = "cpu"
                if os.path.exists(model_cfg_path):
                    try:
                        with open(model_cfg_path, "r", encoding="utf-8") as f:
                            m_cfg = json.load(f)
                            current_device = m_cfg.get("device", current_device)
                    except Exception:
                        pass

                if save_config_to_file(model_name, config_base_prompt, params, device=current_device):
                    console.print("[green]Parameters updated successfully![/green]")
            except ValueError:
                console.print("[red]Invalid input. Operation cancelled.[/red]")

        else:
            console.print("[white]-[white] [red]Unknown command[/red]")

# chat_command_list_function
def chat_command_list_function():
    """Displays commands available in the chat session."""
    print("\n")
    console.print(Rule("Chat", style="cyan"))
    console.print(
        Panel(

            "[bold white]Chat Commands[/bold white]\n"
            "[cyan]Save[/cyan] - save the current chat session.\n"
            "[cyan]Copy[/cyan] - copy the last AI response to clipboard.\n"
            "[cyan]Change[/cyan] - change the current AI model.\n"
            "[red]Exit[/red] - exit the chat session and return to the main command menu.\n"
            "\n"
            "[bold white]Files Commands[/bold white]\n"
            "[cyan]Add[/cyan] - attach a file (.txt, .pdf, .docx, .xlsx, .md, .py, .cpp, img, etc.) for context.\n"
            "[cyan]Remove[/cyan] - remove the attached file.\n"
            "\n"
            "[bold white]Advanced Commands[/bold white]\n"
            "[cyan]Search[/cyan] - search the web with Tavily and send results to the AI.\n"
            "[cyan]Agent[/cyan] - work directly with files using AI.\n",
            border_style="yellow"
        )
    )

    print()

    warning = Text("WARNING: AI models can produce incorrect or misleading responses due to BIAS or Allucinations. For thi reason always check AI repsonse", style="red bold")
    console.print(Align.center(warning))
    print()


def get_memory_path():
    """Returns the path to memory.json in configs."""
    return os.path.join(get_base_path(), "configs", "memory.json")

def load_memories():
    """Loads memories from configs/memory.json."""
    global config_memories
    path = get_memory_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                config_memories = json.load(f)
        except Exception:
            config_memories = []
    else:
        config_memories = []

def save_memories():
    """Saves memories to configs/memory.json."""
    path = get_memory_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(config_memories, f, ensure_ascii=False, indent=2)
    except Exception as e:
        console.print(f"[red]Error saving memories: {e}[/red]")

def memory_add():
    """Add a new memory."""
    global config_memories
    memory = console.input("[cyan]Enter memory (what the AI should remember) > [/cyan]").strip()
    if not memory:
        console.print("[yellow]Cancelled.[/yellow]")
        return
    config_memories.append(memory)
    save_memories()
    console.print("[green]Memory saved![/green]")

def memory_view():
    """View all memories."""
    if not config_memories:
        console.print("[yellow]No memories saved.[/yellow]")
        return
    print()
    table = Table(title="Saved Memories", border_style="cyan")
    table.add_column("#", style="white bold")
    table.add_column("Memory", style="yellow")
    for i, mem in enumerate(config_memories, 1):
        table.add_row(str(i), mem)
    console.print(Align.center(table))

def memory_remove():
    """Remove a specific memory by index."""
    global config_memories
    if not config_memories:
        console.print("[yellow]No memories to remove.[/yellow]")
        return
    memory_view()
    print()
    choice = console.input("[cyan]Enter the number of the memory to remove > [/cyan]").strip()
    if not choice.isdigit():
        console.print("[red]Invalid number.[/red]")
        return
    idx = int(choice) - 1
    if 0 <= idx < len(config_memories):
        removed = config_memories.pop(idx)
        save_memories()
        console.print(f"[green]Memory removed: {removed}[/green]")
    else:
        console.print("[red]Invalid memory number.[/red]")

def memory_clear():
    """Clear all memories."""
    global config_memories
    if not config_memories:
        console.print("[yellow]No memories to clear.[/yellow]")
        return
    confirm = console.input("[red]Are you sure you want to clear all memories? (y/n) > [/red]").strip().lower()
    if confirm == "y":
        config_memories = []
        save_memories()
        console.print("[green]All memories cleared.[/green]")
    else:
        console.print("[yellow]Cancelled.[/yellow]")

# save_chat
def save_chat():
    """Saves the current chat history to a file on Desktop."""
    if not chat_history:
        console.print("[red]No chat history to save.[/red]")
        return

    title = console.input("[cyan]Chat Title > [/cyan]").strip()
    if not title:
        console.print("[red]Title not valid.[/red]")
        return

    safe = "".join(c for c in title if c.isalnum() or c in (' ', '-', '_')).rstrip()
    if not safe:
        safe = "chat"

    chats_dir = get_chats_dir()
    filename = f"{safe}.txt"
    path = os.path.join(chats_dir, filename)

    try:
        with open(path, "w", encoding="utf-8") as f:
            now = datetime.now().strftime("%d/%m/%Y at %H:%M:%S")
            f.write(f"Chat saved on: {now}")
            f.write("\n")
            f.write("=" * 50)
            f.write("\n")
            for line in chat_history:
                f.write(line + "\n")
                f.write("\n")

            console.print(f"[green]Chat saved to '{path}'[/green]")
    except Exception as e:
        console.print(f"[red]Error saving chat: {e}[/red]")

# list_chats_cmd
def list_chats_cmd():
    """Lists all saved chat files in the chats directory."""
    chats_dir = get_chats_dir()
    try:
        chat_files = [f for f in os.listdir(chats_dir) if f.endswith(".txt")]
        if not chat_files:
            console.print("[yellow]No saved chats found.[/yellow]")
            return

        print()
        table = Table(title="Saved Chats", border_style="cyan")
        table.add_column("#", style="white bold")
        table.add_column("Filename", style="yellow")

        for i, fname in enumerate(sorted(chat_files), 1):
            table.add_row(str(i), fname)

        console.print(Align.center(table))
    except Exception as e:
        console.print(f"[red]Error listing chats: {e}[/red]")

# open_chat_cmd
def open_chat_cmd():
    """Displays a saved chat file in the terminal."""
    chats_dir = get_chats_dir()
    chat_name = console.input("[cyan]Enter chat filename (e.g., mychat_01-01-25.txt) > [/cyan]").strip()
    if not chat_name:
        console.print("[yellow]Cancelled.[/yellow]")
        return

    chat_path = os.path.join(chats_dir, chat_name)
    if not os.path.exists(chat_path):
        console.print(f"[red]Chat file '{chat_name}' not found.[/red]")
        return

    try:
        with open(chat_path, "r", encoding="utf-8") as f:
            content = f.read()
        print()
        console.print(Panel(content, title=chat_name, border_style="cyan"))
    except Exception as e:
        console.print(f"[red]Error reading chat: {e}[/red]")

# remove_chat_cmd
def remove_chat_cmd():
    """Deletes a saved chat file."""
    chats_dir = get_chats_dir()
    chat_name = console.input("[cyan]Enter chat filename to delete (e.g., mychat_01-01-25.txt) > [/cyan]").strip()
    if not chat_name:
        console.print("[yellow]Cancelled.[/yellow]")
        return

    chat_path = os.path.join(chats_dir, chat_name)
    if not os.path.exists(chat_path):
        console.print(f"[red]Chat file '{chat_name}' not found.[/red]")
        return

    confirm = console.input(f"[red]Are you sure you want to delete '[bold]{chat_name}[/bold]'? (y/n) > [/red]").strip().lower()
    if confirm == 'y':
        try:
            os.remove(chat_path)
            console.print(f"[green]Chat '[bold]{chat_name}[/bold]' deleted successfully.[/green]")
        except Exception as e:
            console.print(f"[red]Error deleting chat: {e}[/red]")
    else:
        console.print("[yellow]Cancelled.[/yellow]")

# chat_loop
def chat_loop():
    """The main interacting loop for user and AI conversations."""
    global config_model_path, config_base_prompt, attached_file_content, attached_filename
    while True:
        print()
        user_input = console.input("[cyan]You > [/cyan]").strip()

        if user_input.lower() == "save":
            save_chat()
        elif user_input.lower() == "copy":
            last_msg = None
            for msg in reversed(chat_history):
                if " > " in msg and not msg.startswith("USER:"):
                    last_msg = msg.split(" > ", 1)[1] if " > " in msg else msg
                    break
            if last_msg:
                try:
                    if not copy_to_clipboard(last_msg):
                        raise RuntimeError(
                            "Clipboard not available on Windows. Make sure Windows clipboard services are working."
                        )
                    print()
                    console.print("[green]Response copied to clipboard[/green]")
                except Exception as e:
                    console.print(f"[red]Error copying: {e}[/red]")
            else:
                console.print("[yellow]No AI response found to copy.[/yellow]")

        elif user_input.lower() == "add":
            file_path = console.input("[cyan]Insert file path (.txt, .pdf, .docx, .xlsx, .md, .py, .cpp, img, etc.) > [/cyan]").strip()
            supported_extensions = [".txt", ".pdf", ".docx", ".xlsx", ".xls", ".md", ".py", ".cpp", ".c", ".java", ".js", ".html", ".css", ".json", ".xml", ".csv", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"]

            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in supported_extensions:
                print()
                console.print("[red]Error: File type not supported. Supported types: .txt, .pdf, .docx, .xlsx, .md, .py, .cpp, img, etc.[/red]")
            elif not os.path.exists(file_path):
                print()
                console.print("[red]Error: File not found.[/red]")
            else:
                try:
                    if file_path.lower().endswith(".pdf"):
                        reader = pypdf.PdfReader(file_path)
                        text_list = []
                        for page in reader.pages:
                            text_list.append(page.extract_text())
                        attached_file_content = "\n".join(page_text for page_text in text_list if page_text)
                    elif file_path.lower().endswith(".docx"):
                        doc = Document(file_path)
                        text_list = []
                        for paragraph in doc.paragraphs:
                            text_list.append(paragraph.text)
                        attached_file_content = "\n".join(text_list)
                    elif file_path.lower().endswith((".xlsx", ".xls")):
                        workbook = openpyxl.load_workbook(file_path, read_only=True)
                        text_list = []
                        for sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            text_list.append(f"Sheet: {sheet_name}")
                            for row in sheet.iter_rows(values_only=True):
                                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                                text_list.append(row_text)
                        attached_file_content = "\n".join(text_list)
                    elif file_ext in [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"]:
                        attached_file_content = f"[Image file: {os.path.basename(file_path)}]"
                    else:
                        # For text-based files (.txt, .md, code files, etc.)
                        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                            attached_file_content = f.read()
                    print()
                    global attached_filename
                    attached_filename = os.path.basename(file_path)
                    console.print(Align.center(Text(f"File attached: {attached_filename}", style="bold yellow")))
                    chat_history.append(f"User ADD: {attached_filename}")
                except Exception as e:
                    print()
                    console.print(f"[red]Error reading file: {e}[/red]")

        elif user_input.lower() == "change":
            change_model_function()

        elif user_input.lower() == "remove":
            if attached_file_content:
                fname = attached_filename or "file"
                attached_file_content = None
                attached_filename = None
                print()
                console.print(Align.center(Text(f"File {fname} removed", style="bold yellow")))
                chat_history.append(f"User REMOVE: {fname}")
            else:
                print()
                console.print("[red]No file attached to remove.[/red]")

        elif user_input.lower() == "agent":
            agent_mode_function()
            continue 

        elif user_input.lower() == "search":
            if not config_tavily_api_key:
                console.print("[red]Tavily API key not configured. Use More > API Config first.[/red]")
                continue

            search_target = console.input("[cyan]Enter a website, link, or topic to search > [/cyan]").strip()
            if not search_target:
                console.print("[yellow]Cancelled.[/yellow]")
                continue

            user_task = console.input("[cyan]What should the AI do with these results? > [/cyan]").strip()
            if not user_task:
                console.print("[yellow]Cancelled.[/yellow]")
                continue

            try:
                with console.status("[yellow]Searching Tavily web...[/yellow]", spinner="dots"):
                    tavily_payload = tavily_search(search_target, config_tavily_api_key)
                tavily_context = format_tavily_context(tavily_payload)
            except Exception as e:
                console.print(f"[red]Error searching web: {e}[/red]")
                continue

            chat_history.append(f"USER: Command Search\nTarget: {search_target}\nRequest: {user_task}")
            web_prompt = (
                f"{active_base_prompt or DEFAULT_BASE_PROMPT}\n\n"
                "You are given Tavily web search results. Use them to complete the user's request.\n\n"
                f"Search target: {search_target}\n"
                f"User request: {user_task}\n\n"
                f"Tavily context:\n{tavily_context}\n\n"
                "Answer:"
            )

            ai_response = generate_ai_response(web_prompt)
            if ai_response:
                chat_history.append(f"AI respond to search request: {ai_response}")
                typing_effect(ai_response, active_model)

        elif user_input.lower() == "exit":
            console.print(Rule("Chat Terminated", style="cyan"))
            shutdown_ollama_everywhere()
            return

        else:
            if not ollama_ready:
                print()
                console.print("[red]Ollama model not loaded[/red]")
                continue

            chat_history.append(f"User: {user_input}")
            memory_block = ""
            if config_memories:
                memory_lines = "\n".join(f"- {m}" for m in config_memories)
                memory_block = f"\n\nMemory context:\n{memory_lines}"
            if attached_file_content:
                full_prompt = f"{active_base_prompt or DEFAULT_BASE_PROMPT}{memory_block}\n\nAttached file content:\n{attached_file_content}\n\nUser: {user_input}\nAI:"
            else:
                full_prompt = f"{active_base_prompt or DEFAULT_BASE_PROMPT}{memory_block}\n\nUser: {user_input}\nAI:"

            ai_response = generate_ai_response(full_prompt)
            if ai_response:
                typing_effect(ai_response, active_model)

# Start Function
introduction()